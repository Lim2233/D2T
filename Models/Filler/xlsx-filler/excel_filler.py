import os
import glob
import json
import logging
import argparse
from openpyxl import load_workbook

# 配置日志格式
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def fill_table_from_json(worksheet, column_names, json_data, start_row=2):
    """
    根据 JSON 数据填充工作表

    :param worksheet:    openpyxl 工作表对象
    :param column_names: 第一行的列名列表
    :param json_data:    JSON 数据列表，每个元素为字典
    :param start_row:    数据填充的起始行号，默认为 2
    """
    # 建立列索引 → 字段名的映射
    col_key_map = {
        idx: name for idx, name in enumerate(column_names, start=1) if name is not None
    }

    for row_offset, record in enumerate(json_data):
        row_idx = start_row + row_offset
        for col_idx, key in col_key_map.items():
            if key in record:
                worksheet.cell(row=row_idx, column=col_idx, value=record[key])


class ExcelFiller:
    """Excel 数据填充器，支持批处理文件夹内所有 .xlsx 文件"""

    def __init__(
        self,
        json_path,
        input_folder,
        output_folder,
        sheet_name=None,
        header_row=1,
        start_row=2
    ):
        """
        :param json_path:     JSON 数据文件路径
        :param input_folder:  输入 Excel 文件夹路径
        :param output_folder: 输出文件夹路径
        :param sheet_name:    要处理的工作表名称，None 表示使用活动工作表
        :param header_row:    列名所在行号（从 1 开始）
        :param start_row:     数据填充的起始行号
        """
        self.json_path = json_path
        self.input_folder = input_folder
        self.output_folder = output_folder
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.start_row = start_row

        self._json_data = None

    def load_json(self):
        """加载 JSON 数据，并缓存"""
        if self._json_data is None:
            try:
                with open(self.json_path, 'r', encoding='utf-8') as f:
                    self._json_data = json.load(f)
                logger.info(f"成功加载 JSON 数据，共 {len(self._json_data)} 条记录")
            except Exception as e:
                logger.error(f"读取 JSON 文件失败：{e}")
                raise
        return self._json_data

    def process_file(self, file_path):
        """处理单个 Excel 文件"""
        file_name = os.path.basename(file_path)
        output_path = os.path.join(self.output_folder, file_name)

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[self.sheet_name] if self.sheet_name else wb.active

            # 读取指定行的列名
            header_cells = list(ws.iter_rows(min_row=self.header_row, max_row=self.header_row, values_only=True))
            if not header_cells:
                logger.warning(f"文件 {file_name} 第 {self.header_row} 行为空，跳过")
                wb.close()
                return

            column_names = header_cells[0]  # 取第一行（其实只有一行）

            # 获取 JSON 数据（已缓存）
            json_data = self.load_json()

            # 填充数据
            fill_table_from_json(ws, column_names, json_data, start_row=self.start_row)

            # 保存
            wb.save(output_path)
            logger.info(f"已填充并保存：{output_path}")

            wb.close()
        except Exception as e:
            logger.error(f"处理文件 '{file_name}' 时出错：{e}")

    def run(self):
        """执行批处理"""
        # 确保输出文件夹存在
        os.makedirs(self.output_folder, exist_ok=True)

        pattern = os.path.join(self.input_folder, "*.xlsx")
        xlsx_files = glob.glob(pattern)

        if not xlsx_files:
            logger.warning(f"在文件夹 '{self.input_folder}' 中没有找到 .xlsx 文件。")
            return

        logger.info(f"找到 {len(xlsx_files)} 个 Excel 文件，开始处理...")
        for file_path in xlsx_files:
            self.process_file(file_path)
        logger.info("所有文件处理完毕。")


def main():
    """命令行入口"""
    parser = argparse.ArgumentParser(description="根据 JSON 数据批量填充 Excel 模板")
    parser.add_argument("--json", required=True, help="JSON 数据文件路径")
    parser.add_argument("--input", required=True, help="输入 Excel 文件夹路径")
    parser.add_argument("--output", required=True, help="输出文件夹路径")
    parser.add_argument("--sheet", default=None, help="工作表名称，默认使用活动工作表")
    parser.add_argument("--header-row", type=int, default=1, help="列名所在行号，默认 1")
    parser.add_argument("--start-row", type=int, default=2, help="数据填充起始行号，默认 2")

    args = parser.parse_args()

    filler = ExcelFiller(
        json_path=args.json,
        input_folder=args.input,
        output_folder=args.output,
        sheet_name=args.sheet,
        header_row=args.header_row,
        start_row=args.start_row
    )
    filler.run()


if __name__ == "__main__":
    main()